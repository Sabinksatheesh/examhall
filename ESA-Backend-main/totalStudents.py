import openpyxl
import sys
import json

# Get input from command-line argument
data_json = sys.argv[1]
data = json.loads(data_json)

details = data["details"]
sem = details[0].get("sem")

s = set()
for item in details:
    s.add(item.get("slot"))

slot_list = list(s)
noOfStudents = 0

for input_slot in slot_list:
    code_list = [d["branch"] for d in details if d["slot"] == input_slot]
    
    for code in code_list:
        check_supply = 1
        file_path = f'./updatedExcels/S{sem}_{code}.xlsx'
        try:
            wb_branch = openpyxl.load_workbook(file_path)
            ws_branch_reg = wb_branch[input_slot]
            
            # Count normal students
            noOfStudents += ws_branch_reg.max_row

            # Count supply students if sheet exists
            try:
                ws_branch_sply = wb_branch[f"{input_slot}_supply"]
                noOfStudents += ws_branch_sply.max_row
            except KeyError:
                check_supply = 0
        except FileNotFoundError:
            continue  # Skip missing files

# Room calculations
rem = noOfStudents % 30
noOfRooms30 = noOfStudents // 30
noOfRooms60 = 0

if rem <= 10:
    noOfRooms30 -= 1
    noOfRooms60 = 1
elif rem >= 20:
    noOfRooms30 += 1
else:
    noOfRooms30 -= 2
    noOfRooms60 = 2

# Ensure only JSON output is printed
# output_data = {
#     "noOfStudents": noOfStudents,
#     "noOfRooms30": noOfRooms30,
#     "noOfRooms60": noOfRooms60
# }

print(noOfStudents) 
